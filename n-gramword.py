from collections import Counter
import os
import pandas as pd
import openpyxl
from openpyxl import load_workbook
def converter(input_excel, output_text):
    # Read the  Facebook scrapped Excel sheet
    group_excel=pd.read_excel(input_excel,sheet_name='groups')
    group_name=group_excel.iloc[0, group_excel.columns.get_loc('name')]
    group_name=group_name.replace(' ','_').lower()
    #print(group_name)
    my_excel = pd.read_excel(input_excel, sheet_name='posts')

    # Extract column L which has the posts 
    column_L_data = my_excel['text']
    

    # Write all the posts  data to a text file  named as '\\Ngram\\'+file_name+'_'+'unfiltered_list_of_words.txt'
    
    with open(output_text, 'w', encoding='utf-8') as text_file:
        for value in column_L_data:
            text_file.write(str(value) + '\n')
 # Now we split all words into a new line  
def split_words(input_file, output_file):
    with open(input_file, 'r', encoding='utf-8') as infile, open(output_file, 'w', encoding='utf-8') as outfile:
        for line in infile:
            words = line.split()
            for word in words:
                cleanword=word.replace('*', '').replace('#', '').lower()
                outfile.write(cleanword + '\n')

# Merges the words based on the required N value
def merge_lines(file1, output_file, num):
    with open(file1, 'r', encoding='utf-8') as f1, open(output_file, 'w', encoding='utf-8') as out_file:
        lines = [line.strip() for line in f1]

        # Generate lines by concatenating the current line with the next num-1 lines
        for i in range(len(lines) - num + 1):
            merged_line = ' '.join(lines[i:i+num])
            # Write the merged line to the output file
            out_file.write(merged_line + '\n')






def count_word_occurrences(file_path, output_file_path, num,excel_result):
    with open(file_path, 'r', encoding='utf-8') as file:

        lines = [line.strip() for line in file]

    # Generate word sequences and count occurrences
    #word_sequences = [' '.join(lines[i:i+num]) for i in range(len(lines))]

    occurrences = Counter(lines)
    #loading the excel sheets with desired names
    try:
        workbook = openpyxl.load_workbook(excel_result)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
    #workbook=load_workbook(excel_result)
    ngram=str(num)+ ' word occurrences'
    if ngram in workbook.sheetnames:
        
        worksheet = workbook[ngram]
    else:
        
        worksheet = workbook.create_sheet(title=ngram)
    worksheet=workbook[ngram]
    #worksheet.cell(row=1,column=1).value= 'Pattern'
    #worksheet.cell(row=1,column=2).value= 'Repetation'
    header_row = ['Pattern', 'Repetation']  # Add your column names
    worksheet.append(header_row)
    row=2

    # Write occurrences to the output file
    with open(output_file_path, 'w', encoding='utf-8') as output_file:
        for sequence, count in occurrences.most_common():
            output_file.write(f"{sequence} - {count}\n")
            worksheet.cell(row=row, column=1).value = sequence
            worksheet.cell(row=row, column=2).value = count
            row=row+1
    workbook.save(excel_result)



# Example usage with num=3
# Replace 'input_file.xlsx' with the actual name of your Excel file
# Replace 'output_file.txt' with the desired name for your text file
curr_directory=os.getcwd()
file_name='CardersGroups11-142.32am'
inputfile=curr_directory+'/DATA/'+file_name+'.xlsx'
temp_out_file1=curr_directory+'\\Ngram\\'+file_name+'_'+'unfiltered_list_of_words.txt'
converter(inputfile, temp_out_file1)
group_excel=pd.read_excel(inputfile,sheet_name='groups')
group_name=group_excel.iloc[0, group_excel.columns.get_loc('name')]
group_name=group_name.replace(' ','_').lower()
print(group_name)


temp_out_file=curr_directory+'/Ngram/'+file_name+' _'+'unfiltered_list_of_words.txt'
split_words(temp_out_file1, temp_out_file)
 ##cleanword=word.replace('*', '').replace('#', '')  

 
num = int(input(" what is the number of words pattern you are looking for"))
list_of_all_n_grams=curr_directory+'/Ngram/'+file_name+' _'+str(num)+'_list_of_all_n.txt'
merge_lines(temp_out_file,list_of_all_n_grams , num)
# Replace 'input_file.txt' with the actual name of your input text file
# Replace 'output_file.txt' with the desired name for your output text file

output_n_gram=curr_directory+ '/Ngram/'+file_name+' _'+str(num)+'gram.txt'
excel_result=curr_directory+ '/Ngram/Result_Excel/'+group_name+'.xlsx'
count_word_occurrences(list_of_all_n_grams, output_n_gram, num,excel_result)



